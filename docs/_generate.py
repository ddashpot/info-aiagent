#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
docs/_generate.py
既存データから「開発」「リスク管理」の知識を *省略なく・改変なく* Markdown へ体系化する。
出力:
  docs/INDEX.md
  docs/開発体系.md
  docs/リスク管理体系.md
データ自身が持つ分類（domain / category / family / framework / 区分）に沿って一覧化し、
独自の分類は作らない。全項目・全フィールドを出力する。
再実行すれば常に同じ結果になる（恣意性なし）。
"""
import json, os, re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def p(*a): return os.path.join(ROOT, *a)
def load(fn): return json.load(open(p(fn), encoding="utf-8"))

def cell(s):
    """Markdown テーブルセル用エスケープ（内容は変えない・改行は<br>）。"""
    if s is None: return ""
    s = str(s)
    return s.replace("\\", "\\\\").replace("|", "\\|").replace("\n", "<br>")

def lst(xs):
    if not xs: return "—"
    return "、".join(str(x) for x in xs)

OUT = []
def w(line=""): OUT.append(line)
def flush(fn):
    open(p("docs", fn), "w", encoding="utf-8").write("\n".join(OUT) + "\n")
    OUT.clear()

# =========================================================================
# 開発体系.md
# =========================================================================
def gen_dev():
    d = load("ai_agent_catalog_v8-1_data.json")
    items = d["items"]
    agents = [i for i in items if i["domain"] == "agent"]
    tools  = [i for i in items if i["domain"] == "tool"]
    mems   = [i for i in items if i["domain"] == "memory"]

    w("# 開発体系一覧（AIエージェント設計）")
    w()
    w("> **出典**: `ai_agent_catalog_v8-1_data.json` / `SKILL.md` / `AIエージェント基盤比較.xlsx`  ")
    w("> 本書は上記データから機械生成（`docs/_generate.py`）。**要約・省略・改変なし**で全項目・全フィールドを掲載する。")
    w()
    # 0. サマリ
    w("## 0. 件数サマリ")
    w()
    w("| 区分 | 件数 |")
    w("|---|---|")
    w(f"| エージェント構成（agent） | {len(agents)} |")
    w(f"| メモリ設計（memory） | {len(mems)} |")
    w(f"| ツールユース（tool） | {len(tools)} |")
    w(f"| メモリファミリ（families） | {len(d['families'])} |")
    w(f"| 推奨ユースケース構成（recommended） | {len(d['recommended'])} |")
    w(f"| 相性: 構成×メモリ（amMatrix） | {len(d['amMatrix'])} |")
    w(f"| 相性: 構成×ツール（atMatrix） | {len(d['atMatrix'])} |")
    w(f"| タグ（tags） | {len(d['tags'])} |")
    w()
    w("### 目次")
    w("1. [設計パターンカタログ](#1-設計パターンカタログ)")
    w("2. [メモリファミリと共通指針](#2-メモリファミリと共通指針)")
    w("3. [推奨ユースケース構成](#3-推奨ユースケース構成)")
    w("4. [相性マトリクス](#4-相性マトリクス)")
    w("5. [タグ一覧](#5-タグ一覧)")
    w("6. [実装基盤ベンダー比較](#6-実装基盤ベンダー比較)")
    w("7. [参照ドキュメント SKILL.md](#7-参照ドキュメント-skillmd)")
    w()

    def emit_item(x):
        w(f"##### {x['name']}  `{x['id']}`")
        w()
        w("| 属性 | 値 |")
        w("|---|---|")
        w(f"| 分類(category) | {cell(x.get('category'))} |")
        label = {"agent":"配置(placement)","tool":"通信様式(placement)","memory":"保存場所(placement)"}[x["domain"]]
        w(f"| {label} | {cell(x.get('placement'))} |")
        w(f"| メモリ/状態(memory) | {cell(x.get('memory'))} |")
        w(f"| 役割(role) | {cell(x.get('role'))} |")
        if x.get("expectedTools"):
            w(f"| 想定ツール(expectedTools) | {cell(x.get('expectedTools'))} |")
        if x.get("usedByAgents"):
            w(f"| 使う構成(usedByAgents) | {cell(lst(x.get('usedByAgents')))} |")
        if x.get("primaryFamily"):
            w(f"| 主ファミリ(primaryFamily) | {cell(x.get('primaryFamily'))} |")
        if x.get("familyCode"):
            w(f"| ファミリコード(familyCode) | {cell(x.get('familyCode'))} |")
        if x.get("familyCodes"):
            w(f"| ファミリコード(familyCodes) | {cell(lst(x.get('familyCodes')))} |")
        kw = x.get("keywords")
        w(f"| キーワード(keywords) | {cell(lst(kw) if isinstance(kw, list) else kw)} |")
        w()
        w(f"**要約**: {x.get('summary','—')}")
        w()
        # steps
        if x.get("steps"):
            w("**処理ステップ（steps）**")
            w()
            w("| # | 工程(title) | 一言(short) | 説明(description) |")
            w("|---|---|---|---|")
            for n, s in enumerate(x["steps"], 1):
                w(f"| {n} | {cell(s.get('title'))} | {cell(s.get('short'))} | {cell(s.get('description'))} |")
            w()
        # useCases
        if x.get("useCases"):
            w("**ユースケース例（useCases）**")
            w()
            w("| 例(title) | 場面(scene) | 入力(input) | 成功条件(success) |")
            w("|---|---|---|---|")
            for u in x["useCases"]:
                w(f"| {cell(u.get('title'))} | {cell(u.get('scene'))} | {cell(u.get('input'))} | {cell(u.get('success'))} |")
            w()
        # risks
        if x.get("risksControls"):
            w("**リスクと対策（risksControls）**")
            w()
            w("| リスク(risk) | 対策(control) |")
            w("|---|---|")
            for r in x["risksControls"]:
                w(f"| {cell(r.get('risk'))} | {cell(r.get('control'))} |")
            w()
        # memoryUse
        mu = x.get("memoryUse") or {}
        if mu:
            w("**メモリの使い方（memoryUse）**")
            w()
            w(f"- メモリ大分類(families): {cell(lst(mu.get('families') or x.get('memoryFamilies')))}")
            w(f"- 具体パターン(branches): {cell(lst(mu.get('branches') or x.get('memoryBranches')))}")
            w(f"- 主たるメモリ(primary): {cell(mu.get('primary'))}")
            w(f"- 読み出し(read): {cell(mu.get('read'))}")
            w(f"- 書き込み(write): {cell(mu.get('write'))}")
            w()
        # legend
        if x.get("legend"):
            w("**図の凡例（legend）**")
            w()
            w("| 記号(symbol) | 意味(meaning) |")
            w("|---|---|")
            for g in x["legend"]:
                w(f"| {cell(g.get('symbol'))} | {cell(g.get('meaning'))} |")
            w()
        # diagrams
        for key, title in [("dataFlowMermaid","データフロー図"),
                           ("architectureMermaid","構成図"),
                           ("memoryFlowMermaid","メモリ使用図")]:
            if x.get(key):
                w(f"**{title}（{key}）**")
                w()
                w("```mermaid")
                w(x[key])
                w("```")
                w()
        w("---")
        w()

    w("## 1. 設計パターンカタログ")
    w()
    w("### 1.1 エージェント構成（agent）")
    w()
    # group by category, preserving data order
    cats = []
    for a in agents:
        if a["category"] not in cats: cats.append(a["category"])
    for c in cats:
        w(f"#### {c}")
        w()
        group = [a for a in agents if a["category"] == c]
        w("| ID | 名称 | 配置 | メモリ | 役割 |")
        w("|---|---|---|---|---|")
        for a in group:
            w(f"| `{a['id']}` | {cell(a['name'])} | {cell(a.get('placement'))} | {cell(a.get('memory'))} | {cell(a.get('role'))} |")
        w()
        for a in group:
            emit_item(a)

    w("### 1.2 メモリ設計（memory）")
    w()
    w("| ID | 名称 | 保存場所 | 役割 | 主ファミリ |")
    w("|---|---|---|---|---|")
    for m in mems:
        w(f"| `{m['id']}` | {cell(m['name'])} | {cell(m.get('placement'))} | {cell(m.get('role'))} | {cell(m.get('primaryFamily'))} |")
    w()
    for m in mems:
        emit_item(m)

    w("### 1.3 ツールユース（tool）")
    w()
    w("| ID | 名称 | 通信様式 | 役割 |")
    w("|---|---|---|---|")
    for t in tools:
        w(f"| `{t['id']}` | {cell(t['name'])} | {cell(t.get('placement'))} | {cell(t.get('role'))} |")
    w()
    for t in tools:
        emit_item(t)

    # 2. families
    w("## 2. メモリファミリと共通指針")
    w()
    w("### 2.1 メモリファミリ（families）")
    w()
    fam = d["families"]
    # families values may be dict; dump fields faithfully
    keys = set()
    for v in fam.values():
        if isinstance(v, dict): keys |= set(v.keys())
    keys = list(keys)
    if keys:
        w("| コード | " + " | ".join(keys) + " |")
        w("|---|" + "|".join(["---"]*len(keys)) + "|")
        for code, v in fam.items():
            if isinstance(v, dict):
                w(f"| `{code}` | " + " | ".join(cell(v.get(k)) for k in keys) + " |")
            else:
                w(f"| `{code}` | {cell(v)} |")
    else:
        for code, v in fam.items():
            w(f"- `{code}`: {cell(v)}")
    w()
    # memoryPrimary mapping
    w("### 2.2 メモリ種別の主ファミリ対応（memoryPrimary）")
    w()
    w("| メモリ種別ID | 主ファミリ |")
    w("|---|---|")
    for k, v in d["memoryPrimary"].items():
        w(f"| `{k}` | {cell(v if isinstance(v,str) else json.dumps(v,ensure_ascii=False))} |")
    w()
    # memoryCommon
    w("### 2.3 メモリ共通指針（memoryCommon）")
    w()
    for k, v in d["memoryCommon"].items():
        w(f"**{k}**")
        w()
        if isinstance(v, list):
            for e in v: w(f"- {cell(e) if not isinstance(e,(dict,list)) else cell(json.dumps(e,ensure_ascii=False))}")
        elif isinstance(v, dict):
            for kk, vv in v.items(): w(f"- {kk}: {cell(vv if not isinstance(vv,(dict,list)) else json.dumps(vv,ensure_ascii=False))}")
        else:
            w(cell(v))
        w()

    # 3. recommended
    w("## 3. 推奨ユースケース構成")
    w()
    w("`recommended`: ユースケースごとの推奨 A（構成）／M（メモリ）／T（ツール）の組み合わせと、その固有リスク。")
    w()
    w("| ID | ユースケース | A 構成 | M メモリ | T ツール | タグ | 要約 |")
    w("|---|---|---|---|---|---|---|")
    for r in d["recommended"]:
        w(f"| {r.get('id')} | {cell(r.get('usecase'))} | `{r.get('A')}` | `{r.get('M')}` | `{r.get('T')}` | {cell(lst(r.get('tags')))} | {cell(r.get('summary'))} |")
    w()
    w("### 3.1 推奨構成ごとの固有リスクと対策")
    w()
    for r in d["recommended"]:
        if r.get("risksControls"):
            w(f"#### {r.get('id')} {r.get('usecase')}（A=`{r.get('A')}` / M=`{r.get('M')}` / T=`{r.get('T')}`）")
            w()
            w("| リスク | 対策 |")
            w("|---|---|")
            for rc in r["risksControls"]:
                w(f"| {cell(rc.get('risk'))} | {cell(rc.get('control'))} |")
            w()

    # 4. matrices
    def emit_matrix(matrix, title, leftlabel, toplabel):
        w(f"### {title}")
        w()
        # collect rows/cols preserving order of appearance
        rows, colsk = [], []
        for k in matrix:
            a, b = k.split("|", 1)
            if a not in rows: rows.append(a)
            if b not in colsk: colsk.append(b)
        w(f"行 = {leftlabel}（{len(rows)}）／ 列 = {toplabel}（{len(colsk)}）。セル = レベル（注記）。")
        w()
        w("| " + leftlabel + " \\ " + toplabel + " | " + " | ".join(f"`{c}`" for c in colsk) + " |")
        w("|---|" + "|".join(["---"]*len(colsk)) + "|")
        for rk in rows:
            line = [f"`{rk}`"]
            for ck in colsk:
                v = matrix.get(f"{rk}|{ck}")
                if v: line.append(cell(f"{v.get('level','')}（{v.get('note','')}）"))
                else: line.append("")
            w("| " + " | ".join(line) + " |")
        w()
    w("## 4. 相性マトリクス")
    w()
    emit_matrix(d["amMatrix"], "4.1 構成 × メモリ（amMatrix）", "構成", "メモリ")
    emit_matrix(d["atMatrix"], "4.2 構成 × ツール（atMatrix）", "構成", "ツール")

    # 5. tags
    w("## 5. タグ一覧")
    w()
    w("`tags`: " + "、".join(d["tags"]))
    w()

    # 6. vendor comparison
    w("## 6. 実装基盤ベンダー比較")
    w()
    w("> 出典: `AIエージェント基盤比較.xlsx`（全行をそのまま転記）")
    w()
    wb = openpyxl.load_workbook(p("AIエージェント基盤比較.xlsx"), data_only=True)
    for ws in wb.worksheets:
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in r)]
        if not rows: continue
        w(f"### シート: {ws.title}")
        w()
        header = rows[0]
        ncol = len(header)
        w("| " + " | ".join(cell(h) for h in header) + " |")
        w("|" + "|".join(["---"]*ncol) + "|")
        for r in rows[1:]:
            r = list(r) + [""]*(ncol-len(r))
            w("| " + " | ".join(cell(c) for c in r[:ncol]) + " |")
        w()

    # 7. SKILL.md outline
    w("## 7. 参照ドキュメント SKILL.md")
    w()
    w("> `SKILL.md` は20種の構成パターンを散文で詳説する一次ドキュメント。以下は見出し一覧（全文は原典参照）。")
    w()
    for line in open(p("SKILL.md"), encoding="utf-8"):
        m = re.match(r"^(#{1,4}) (.+)$", line.rstrip("\n"))
        if m:
            indent = "  " * (len(m.group(1)) - 1)
            w(f"{indent}- {m.group(2)}")
    w()
    flush("開発体系.md")


# =========================================================================
# リスク管理体系.md
# =========================================================================
def gen_risk():
    d = load("ai_security_catalog.json")
    vs = d["vulnerabilities"]
    sev_label = {s["id"]: s["label"] for s in d["severity_scale"]}
    ct_label = {c["id"]: c["label"] for c in d["control_types"]}

    w("# リスク管理体系一覧（AI / エージェンティックAI セキュリティ）")
    w()
    w("> **出典**: `ai_security_catalog.json` / `tools/aiva`（`probes.json`・`tool_registry.json`） / `SKILL_SECURITY.md` / `AIエージェント_ID・セキュリティ要件_4階層_雛形.xlsx`  ")
    w("> 本書は上記データから機械生成（`docs/_generate.py`）。**要約・省略・改変なし**で全項目・全フィールドを掲載する。")
    w()
    meta = d.get("meta", {})
    w(f"> カタログ: **{meta.get('name','')}** v{meta.get('version','')}（更新 {meta.get('updated','')}）")
    w()

    from collections import Counter
    bycat = Counter(v["category"] for v in vs)
    bysev = Counter(v["severity"] for v in vs)
    bysurf = Counter(v.get("surface") for v in vs)
    byfail = Counter(v.get("failure_mode") for v in vs)
    w("## 0. 件数サマリ")
    w()
    w("| 区分 | 件数 |")
    w("|---|---|")
    w(f"| 脆弱性（vulnerabilities） | {len(vs)} |")
    w(f"| 大分類（categories） | {len(d['categories'])} |")
    w(f"| コントロール種別（control_types） | {len(d['control_types'])} |")
    w(f"| 深刻度（severity_scale） | {len(d['severity_scale'])} |")
    w()
    w("### 目次")
    w("1. [脆弱性カタログ](#1-脆弱性カタログ)")
    w("2. [分類軸](#2-分類軸)")
    w("3. [準拠フレームワーク](#3-準拠フレームワーク)")
    w("4. [方法論](#4-方法論)")
    w("5. [検査スキャナ aiva](#5-検査スキャナ-aiva)")
    w("6. [ID・セキュリティ要件 4階層](#6-idセキュリティ要件-4階層)")
    w("7. [参照ドキュメント SKILL_SECURITY.md](#7-参照ドキュメント-skill_securitymd)")
    w()

    # 1. vulnerabilities by category
    w("## 1. 脆弱性カタログ")
    w()
    cats = d["categories"]
    catname = {c["id"]: c["name"] for c in cats}
    # index table (all 34)
    w("### 1.0 一覧（全34件）")
    w()
    w("| ID | 名称 | 深刻度 | 大分類 | 影響サーフェス | 失敗モード |")
    w("|---|---|---|---|---|---|")
    for v in vs:
        w(f"| `{v['id']}` | {cell(v['name'])} | {cell(sev_label.get(v['severity'],v['severity']))} | {cell(catname.get(v['category'],v['category']))} | {cell(v.get('surface'))} | {cell(v.get('failure_mode'))} |")
    w()

    order = [c["id"] for c in cats]
    for cid in order:
        c = next(x for x in cats if x["id"] == cid)
        group = [v for v in vs if v["category"] == cid]
        if not group: continue
        w(f"### 1.{order.index(cid)+1} {c['name']}（{cid}・{len(group)}件）")
        w()
        w(f"> {c.get('summary','')}（フレームワーク: {c.get('framework','')}）")
        w()
        for v in group:
            w(f"#### {v['id']} {v['name']}")
            w()
            w("| 属性 | 値 |")
            w("|---|---|")
            w(f"| 別名(aka) | {cell(lst(v.get('aka')))} |")
            w(f"| 深刻度(severity) | {cell(sev_label.get(v['severity'],v['severity']))} |")
            w(f"| 影響サーフェス(surface) | {cell(v.get('surface'))} |")
            w(f"| 失敗モード(failure_mode) | {cell(v.get('failure_mode'))} |")
            w()
            w(f"**要約(summary)**: {v.get('summary','—')}")
            w()
            w(f"**説明(description)**: {v.get('description','—')}")
            w()
            if v.get("attack_vectors"):
                w("**攻撃ベクトル(attack_vectors)**")
                for a in v["attack_vectors"]: w(f"- {a}")
                w()
            if v.get("affected_patterns"):
                w(f"**影響を受けやすい構成(affected_patterns)**: {lst('`%s`'%x for x in v['affected_patterns'])}")
                w()
            if v.get("controls"):
                w("**コントロール(controls)**")
                w()
                w("| ID | 種別 | 名称 | 説明 | 実装基盤(foundation) |")
                w("|---|---|---|---|---|")
                for ctl in v["controls"]:
                    w(f"| `{ctl.get('id','')}` | {cell(ct_label.get(ctl.get('type'),ctl.get('type')))} | {cell(ctl.get('name'))} | {cell(ctl.get('desc'))} | {cell(ctl.get('foundation'))} |")
                w()
            # framework mappings
            fm = []
            if v.get("atlas"): fm.append(f"MITRE ATLAS: {lst(v['atlas'])}")
            if v.get("nist_rmf"): fm.append(f"NIST AI RMF: {lst(v['nist_rmf'])}")
            if fm:
                w("**フレームワーク対応**")
                for x in fm: w(f"- {x}")
                w()
            if v.get("probes"):
                w(f"**検査プローブ(probes)**: {lst('`%s`'%x for x in v['probes'])}")
                w()
            if v.get("signals"):
                w("**検知シグナル(signals)**")
                for s in v["signals"]: w(f"- {s}")
                w()
            if v.get("references"):
                w("**参考(references)**")
                for ref in v["references"]:
                    if isinstance(ref, dict):
                        w(f"- [{ref.get('title',ref.get('url',''))}]({ref.get('url','')})")
                    else:
                        w(f"- {ref}")
                w()
            w("---")
            w()

    # 2. 分類軸
    w("## 2. 分類軸")
    w()
    w("### 2.1 大分類（categories）")
    w()
    w("| ID | 名称 | フレームワーク | 概要 |")
    w("|---|---|---|---|")
    for c in cats:
        w(f"| `{c['id']}` | {cell(c['name'])} | {cell(c.get('framework'))} | {cell(c.get('summary'))} |")
    w()
    w("### 2.2 深刻度スケール（severity_scale）")
    w()
    w("| ID | ラベル | 重み(weight) |")
    w("|---|---|---|")
    for s in d["severity_scale"]:
        w(f"| `{s['id']}` | {cell(s['label'])} | {s.get('weight')} |")
    w()
    w("### 2.3 コントロール種別（control_types）")
    w()
    w("| ID | ラベル |")
    w("|---|---|")
    for c in d["control_types"]:
        w(f"| `{c['id']}` | {cell(c['label'])} |")
    w()
    w("### 2.4 影響サーフェス分布（surface）")
    w()
    w("| サーフェス | 件数 |")
    w("|---|---|")
    for k, n in bysurf.most_common():
        w(f"| {cell(k)} | {n} |")
    w()
    w("### 2.5 失敗モード分布（failure_mode）")
    w()
    w("| 失敗モード | 件数 |")
    w("|---|---|")
    for k, n in byfail.most_common():
        w(f"| {cell(k)} | {n} |")
    w()

    # 3. frameworks
    w("## 3. 準拠フレームワーク")
    w()
    w("| ID | 名称 | 参照 |")
    w("|---|---|---|")
    for fw in meta.get("frameworks", []):
        w(f"| `{fw.get('id')}` | {cell(fw.get('name'))} | {cell(fw.get('ref'))} |")
    w()

    # 4. methodology
    w("## 4. 方法論（methodology）")
    w()
    def dump(node, depth=0):
        ind = "  " * depth
        if isinstance(node, dict):
            for k, v in node.items():
                if isinstance(v, (dict, list)):
                    w(f"{ind}- **{k}**:")
                    dump(v, depth+1)
                else:
                    w(f"{ind}- **{k}**: {cell(v)}")
        elif isinstance(node, list):
            for e in node:
                if isinstance(e, dict):
                    # render dict on one bullet
                    parts = []
                    for k, v in e.items():
                        parts.append(f"{k}={cell(v) if not isinstance(v,(dict,list)) else json.dumps(v,ensure_ascii=False)}")
                    w(f"{ind}- " + " / ".join(parts))
                else:
                    w(f"{ind}- {cell(e)}")
        else:
            w(f"{ind}- {cell(node)}")
    dump(d["methodology"])
    w()

    # 5. aiva scanner
    probes = load("tools/aiva/aiva/data/probes.json")
    treg = load("tools/aiva/aiva/data/tool_registry.json")
    w("## 5. 検査スキャナ aiva")
    w()
    w(f"> 出典: `tools/aiva/aiva/data/probes.json`（{len(probes['probes'])}プローブ） / `tool_registry.json`（ツール{len(treg['tools'])}・防御{len(treg['defenses'])}・種別{len(treg['kinds'])}）")
    w()
    w("### 5.1 検査プローブ（probes）")
    w()
    w("| ID | 対象脆弱性 | 大分類 | 深刻度 | モード | 期待 | 手法(technique) | タイトル | タグ |")
    w("|---|---|---|---|---|---|---|---|---|")
    for pr in probes["probes"]:
        w(f"| `{pr.get('id')}` | `{pr.get('vuln')}` | {cell(pr.get('category'))} | {cell(pr.get('severity'))} | {cell(pr.get('mode'))} | {cell(pr.get('expect'))} | {cell(pr.get('technique'))} | {cell(pr.get('title'))} | {cell(lst(pr.get('tags')))} |")
    w()
    w("### 5.2 外部ツール種別（kinds）")
    w()
    w("| ID | 名称 |")
    w("|---|---|")
    for k in treg["kinds"]:
        w(f"| `{k.get('id')}` | {cell(k.get('name'))} |")
    w()
    w("### 5.3 連携ツール（tools）")
    w()
    w("| ID | 名称 | ベンダー | 種別 | カバー脆弱性(covers) | インストール | ホームページ |")
    w("|---|---|---|---|---|---|---|")
    for t in treg["tools"]:
        w(f"| `{t.get('id')}` | {cell(t.get('name'))} | {cell(t.get('vendor'))} | {cell(t.get('kind'))} | {cell(lst(t.get('covers')))} | {cell(t.get('install'))} | {cell(t.get('homepage'))} |")
    w()
    w("### 5.4 防御実装（defenses）")
    w()
    w("| ID | 名称 | ベンダー | 種別 | 実装コントロール(implements) | ホームページ |")
    w("|---|---|---|---|---|---|")
    for df in treg["defenses"]:
        w(f"| `{df.get('id')}` | {cell(df.get('name'))} | {cell(df.get('vendor'))} | {cell(df.get('kind'))} | {cell(lst(df.get('implements')))} | {cell(df.get('homepage'))} |")
    w()

    # 6. ID・security 4-layer xlsx
    w("## 6. ID・セキュリティ要件 4階層")
    w()
    w("> 出典: `AIエージェント_ID・セキュリティ要件_4階層_雛形.xlsx`（全行をそのまま転記）。`区分` 列が **開発 / 管理** を示す。")
    w()
    wb = openpyxl.load_workbook(p("AIエージェント_ID・セキュリティ要件_4階層_雛形.xlsx"), data_only=True)
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in r)]
        if not rows: continue
        w(f"### シート: {ws.title}")
        w()
        # find header row (the one containing 'L3' or 'L2' or '区分')
        hidx = 0
        for i, r in enumerate(rows):
            joined = " ".join(str(c) for c in r if c)
            if ("L3" in joined or "区分" in joined) and len([c for c in r if c and str(c).strip()]) >= 3:
                hidx = i; break
        # emit any preamble rows above header as notes
        for r in rows[:hidx]:
            txt = " ".join(str(c) for c in r if c and str(c).strip())
            if txt: w(f"> {cell(txt)}")
        if hidx: w()
        header = rows[hidx]
        ncol = len(header)
        w("| " + " | ".join(cell(h) for h in header) + " |")
        w("|" + "|".join(["---"]*ncol) + "|")
        for r in rows[hidx+1:]:
            r = list(r) + [""]*(ncol-len(r))
            w("| " + " | ".join(cell(c) for c in r[:ncol]) + " |")
        w()

    # 7. SKILL_SECURITY.md outline
    w("## 7. 参照ドキュメント SKILL_SECURITY.md")
    w()
    w("> 評価スイートの運用・データ構造の一次ドキュメント。以下は見出し一覧（全文は原典参照）。")
    w()
    for line in open(p("SKILL_SECURITY.md"), encoding="utf-8"):
        m = re.match(r"^(#{1,4}) (.+)$", line.rstrip("\n"))
        if m:
            indent = "  " * (len(m.group(1)) - 1)
            w(f"{indent}- {m.group(2)}")
    w()
    flush("リスク管理体系.md")


# =========================================================================
# INDEX.md
# =========================================================================
def gen_index():
    dev = load("ai_agent_catalog_v8-1_data.json")
    sec = load("ai_security_catalog.json")
    agents = [i for i in dev["items"] if i["domain"]=="agent"]
    tools  = [i for i in dev["items"] if i["domain"]=="tool"]
    mems   = [i for i in dev["items"] if i["domain"]=="memory"]
    probes = load("tools/aiva/aiva/data/probes.json")
    treg = load("tools/aiva/aiva/data/tool_registry.json")

    w("# 知識ベース体系インデックス")
    w()
    w("AIエージェントの **開発** と **リスク管理** の知識を、既存データから漏れなく一覧化した体系ドキュメント群。")
    w("全ドキュメントは `docs/_generate.py` で再生成可能（要約・省略・改変なし）。")
    w()
    w("## ドキュメント")
    w()
    w("| ドキュメント | 内容 |")
    w("|---|---|")
    w("| [開発体系.md](開発体系.md) | 設計パターン・メモリ・ツール・推奨構成・相性・実装基盤比較 |")
    w("| [リスク管理体系.md](リスク管理体系.md) | 脆弱性カタログ・分類軸・フレームワーク・方法論・スキャナ・ID要件 |")
    w("| [フレームワーク網羅マップ.md](フレームワーク網羅マップ.md) | 各フレームワーク（OWASP/ATLAS/NIST/ISO/CISA/CSA）公式項目 × カタログのカバレッジ |")
    w("| [開発分類_MECE分析.md](開発分類_MECE分析.md) | 開発カタログの分類軸（配置/メモリ/役割）のMECE分析と再設計提案 |")
    w()
    w("## 知識源（一次ソース）")
    w()
    w("### 開発")
    w("| ソース | 内容 | 規模 |")
    w("|---|---|---|")
    w(f"| `ai_agent_catalog_v8-1_data.json` | 設計パターン/メモリ/ツール/推奨/相性/タグ | 構成{len(agents)}・メモリ{len(mems)}・ツール{len(tools)}・推奨{len(dev['recommended'])}・相性{len(dev['amMatrix'])+len(dev['atMatrix'])}・タグ{len(dev['tags'])} |")
    w("| `SKILL.md` | 20構成パターンの散文詳説・HTML出力仕様 | 20パターン |")
    w("| `AIエージェント基盤比較.xlsx` | 実装基盤のベンダー横断比較 | 9ベンダー |")
    w()
    w("### リスク管理")
    w("| ソース | 内容 | 規模 |")
    w("|---|---|---|")
    w(f"| `ai_security_catalog.json` | 脆弱性×コントロール×フレームワーク×プローブ | 脆弱性{len(sec['vulnerabilities'])}・大分類{len(sec['categories'])} |")
    w(f"| `tools/aiva/aiva/data/probes.json` | 検査プローブ | {len(probes['probes'])}プローブ |")
    w(f"| `tools/aiva/aiva/data/tool_registry.json` | 外部ツール/防御実装 | ツール{len(treg['tools'])}・防御{len(treg['defenses'])} |")
    w("| `SKILL_SECURITY.md` | 評価スイート運用・データ構造 | — |")
    w("| `AIエージェント_ID・セキュリティ要件_4階層_雛形.xlsx` | ID/セキュリティ要件（開発/管理区分）×ベンダー | 4階層 |")
    w()
    w("## 再生成")
    w()
    w("```bash")
    w("pip install openpyxl")
    w("python3 docs/_generate.py")
    w("```")
    w()
    flush("INDEX.md")


if __name__ == "__main__":
    os.makedirs(p("docs"), exist_ok=True)
    gen_dev()
    gen_risk()
    gen_index()
    print("generated: docs/INDEX.md, docs/開発体系.md, docs/リスク管理体系.md")
